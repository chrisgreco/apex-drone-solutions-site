import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

targets = []

def add(company, type_, location, email, phone, website, source, priority, notes, subject, body):
    targets.append({
        "Company": company, "Type": type_, "Location": location,
        "Email": email, "Phone": phone, "Website": website,
        "Source": source, "Priority": priority, "Notes": notes,
        "Email Subject": subject, "Email Body": body,
    })

# ── Email templates ──

def acq_subj(name): return f"Partnership interest in {name}"
def acq_body(name): return f"""Hi,

We're building an agricultural drone spraying service in South Jersey and came across {name}.

Rather than a full sale upfront, we're interested in investing alongside an owner with a path to eventually taking over when they're ready to step back. The goal would be to support growth while keeping what's already working (your team, brand, and customer relationships).

I know this is out of the blue, but would you be open to a quick 10-15 minute intro call sometime?

If now's not the right time, I'd be happy to stay in touch down the road.

Best,
Chris"""

def comp_subj(): return "Quick question about aerial application in South Jersey"
def comp_body(name): return f"""Hi,

I'm starting an agricultural drone spraying service focused on specialty crops in South Jersey -- blueberries, cranberries, tomatoes, and the like.

I came across {name} and wanted to reach out. We're not looking to compete on your core business -- we're focused on the acres that are hardest to reach: wet fields, tight rows, buffer zones, and spot treatments where ground rigs and planes aren't ideal.

Would you be open to a brief conversation? I'd love to understand how you see drone spraying fitting alongside traditional aerial application in this market. There may be partnership opportunities worth exploring.

Best,
Chris"""

def pilot_subj(): return "Ag drone spraying opportunity in South Jersey"
def pilot_body(name): return f"""Hi,

I'm launching an agricultural drone spraying service in South Jersey and looking for experienced drone operators who might be interested in flying with us -- either as a pilot, partner, or subcontractor.

We're focused on precision spraying for blueberries, cranberries, and vegetables in Burlington and Cumberland counties. We'll be running a DJI T25 fleet, FAA Part 137 certified, with seasonal contracts already in the pipeline.

If ag drone work interests you, I'd love to chat. Even if the timing isn't right, happy to stay connected.

Best,
Chris"""

def supply_subj(): return "Referral partnership -- drone spraying for your grower customers"
def supply_body(name): return f"""Hi,

I'm starting an agricultural drone spraying service in South Jersey, and I think there's a natural referral partnership between what you do at {name} and what we offer.

Your customers buy the chemicals -- we apply them. When a grower needs a spray application and their ground rig can't get into a wet field, or they need precision coverage in tight blueberry rows, that's where we come in.

Would you be open to a quick conversation about how we might work together? We could send referrals your way too.

Best,
Chris"""

def farm_subj(name): return f"Drone spraying services for {name}"
def farm_body(name, county): return f"""Hi,

I'm launching a precision drone spraying service in South Jersey and reaching out to farms in {county} to gauge interest.

We offer GPS-guided crop spraying at $15-16/acre -- competitive with airplane rates but with better precision, zero soil compaction, and same-day GPS coverage maps showing exactly what was sprayed. We're especially useful for wet fields, tight rows, and hard-to-reach spots.

Would it be helpful to see how drone spraying could work on your operation? Happy to do a quick demo or just chat about your spray needs this season.

Best,
Chris"""

def ext_subj(): return "New ag drone spraying service -- introduction"
def ext_body(): return """Hi,

I'm launching a professional agricultural drone spraying service in South Jersey, focused on precision application for specialty crops -- blueberries, cranberries, tomatoes, peppers, and more.

We'll be FAA Part 137 certified, NJ pesticide applicator licensed, and commercially insured. I'd love to connect with your office to understand what crop protection challenges growers in your county are facing, and to learn how we can be a resource for the farming community.

Would you have 15 minutes for a quick intro call or coffee? I'm also happy to do a demonstration at an upcoming grower event if that would be useful.

Best,
Chris"""

def agctr_subj(): return "Drone spraying demo at Burlington County Ag Center?"
def agctr_body(): return """Hi,

I'm launching a precision agricultural drone spraying service in South Jersey and would love to explore the possibility of hosting a demonstration at the Agricultural Center.

We use DJI T25 drones for GPS-guided crop spraying -- fungicides, pesticides, herbicides -- with zero soil compaction and detailed coverage reports. It's a great fit for South Jersey's specialty crops, especially blueberries and cranberries.

Would you be open to discussing a demo day or presentation at an upcoming grower event? I think local farmers would find it valuable to see the technology in action.

Best,
Chris"""

def get_county(loc):
    if "(" in loc:
        return loc.split("(")[1].replace(")", "")
    return "the area"

# TIER 1: HIGH
add("Influential Drones", "Ag drone spraying (Part 137)", "Marlton, NJ (Burlington Co)", "info@influentialdrones.com", "(856) 281-7545", "influentialdrones.com", "Apify + original", "HIGH", "Only confirmed NJ Part 137 drone sprayer. Top acquisition target.", acq_subj("Influential Drones"), acq_body("Influential Drones"))
add("Alpha Drones USA", "Ag/commercial drone services", "Multi-state (NJ office)", "info@alphadronesusa.com", "1-800-270-4987", "alphadronesusa.com", "Perplexity + scraped", "HIGH", "Active NJ ag drone provider. Competitor or acquisition.", acq_subj("Alpha Drones USA"), acq_body("Alpha Drones USA"))
add("Wings Aerial Applicators", "Traditional aerial applicator", "New Jersey", "wingsaerial@gmail.com", "(609) 760-5653", "wingsaerialapplicators.com", "Perplexity", "HIGH", "Incumbent manned aerial. 18+ years.", comp_subj(), comp_body("Wings Aerial Applicators"))
add("Downstown Aero", "Traditional aerial applicator", "Bridgeton, NJ (Cumberland Co)", "FireCats1@aol.com", "(856) 697-3300", "", "Perplexity", "HIGH", "Founded 1945. Fleet of 13 Ag-Cats.", comp_subj(), comp_body("Downstown Aero"))
add("Burlington County Agricultural Center", "Farmers market / ag center", "Moorestown, NJ (Burlington Co)", "farmmarket@co.burlington.nj.us", "(856) 642-3850", "burlcoagcenter.com", "Apify", "HIGH", "County ag center. Demo venue.", agctr_subj(), agctr_body())
add("Rutgers Extension - Burlington County", "University extension", "Westampton, NJ", "bamka@njaes.rutgers.edu", "(609) 265-5050", "burlington.njaes.rutgers.edu", "Perplexity", "HIGH", "Gateway to Burlington Co blueberry farmers.", ext_subj(), ext_body())

# TIER 2: MEDIUM - Competitors & Adjacent
add("AcuSpray", "Drone spraying technology", "National (NJ service)", "info@acuspray.com", "", "acuspray.com", "Perplexity + scraped", "MEDIUM", "Precision drone spraying. Competitor or partner.", acq_subj("AcuSpray"), acq_body("AcuSpray"))
add("Delaware Valley Spray Services", "Lawn/ag spray service", "Hainesport, NJ (Burlington Co)", "info@delawarevalleysprayservice.com", "(609) 261-9400", "delawarevalleysprayservice.com", "Apify", "MEDIUM", "Spray services in Burlington Co.", comp_subj(), comp_body("Delaware Valley Spray Services"))
add("U.S.A.R. Drone Team (Agra Spray)", "Dealer / ag spray demos", "NJ/Northeast", "usardroneteam@gmail.com", "(732) 245-4234", "usardroneteam.org", "Original + scraped", "MEDIUM", "Regional ag spray operator/dealer.", pilot_subj(), pilot_body("U.S.A.R. Drone Team"))
add("NuWay Ag", "Dealer/pilot network", "New Jersey", "sales@nuwayag.com", "", "nuwayag.com", "Original + scraped", "MEDIUM", "NJ pilot map. Key sourcing channel.", pilot_subj(), pilot_body("NuWay Ag"))

# TIER 2: MEDIUM - Drone operators
add("South Jersey Drone Services", "General drone services", "Sewell, NJ", "info@southjerseydroneservices.com", "(856) 807-1300", "southjerseydroneservices.com", "Original", "MEDIUM", "Local drone ops. Talent source.", pilot_subj(), pilot_body("South Jersey Drone Services"))
add("National Drone Works", "Drone service", "Burlington County, NJ", "info@nationaldroneworks.com", "(856) 685-4188", "nationaldroneworks.com", "Apify", "MEDIUM", "Burlington County drone service.", pilot_subj(), pilot_body("National Drone Works"))
add("The Drone Life, LLC", "Aerial photography/drone", "Medford Lakes, NJ (Burlington Co)", "info@thedronelifenj.com", "(609) 500-5774", "thedronelifenj.com", "Apify", "MEDIUM", "Local drone operator. Potential hire.", pilot_subj(), pilot_body("The Drone Life"))

# TIER 2: MEDIUM - Farm supply
add("Sussex Irrigation & Farm Supply", "Farm equipment supplier", "Bridgeton, NJ (Cumberland Co)", "info@sussexirrigation.com", "(856) 451-1368", "sussexirrigation.com", "Apify", "MEDIUM", "Farm supply. Referral partner.", supply_subj(), supply_body("Sussex Irrigation"))
add("Burlington Agway", "Farm/feed supply", "Burlington, NJ", "burlingtonagway@live.com", "(609) 386-0500", "burlingtonagway.com", "Apify", "MEDIUM", "Feed/farm supply. Referral channel.", supply_subj(), supply_body("Burlington Agway"))
add("Ware's Farm Supply", "Farm supply", "Tabernacle, NJ (Burlington Co)", "waresfarmsupply@gmail.com", "(609) 268-1191", "waresfarmsupply.com", "Apify", "MEDIUM", "Heart of blueberry country.", supply_subj(), supply_body("Ware's Farm Supply"))

# TIER 2: MEDIUM - Farm customers
add("Eastern Fresh Growers", "Vegetable wholesaler", "Cedarville, NJ (Cumberland Co)", "sales@easternfreshgrowers.com", "(856) 447-3563", "easternfreshgrowers.com", "Apify", "MEDIUM", "Major vegetable operation.", farm_subj("Eastern Fresh Growers"), farm_body("Eastern Fresh Growers", "Cumberland County"))
add("Oakland Farms Crop Services", "Farm / crop services", "Bridgeton, NJ (Cumberland Co)", "oaklandfarmsnj@gmail.com", "(856) 451-8224", "", "Apify", "MEDIUM", "Farm with crop services.", farm_subj("Oakland Farms"), farm_body("Oakland Farms", "Cumberland County"))
add("USDA Service Center - Vineland", "Federal ag office", "Vineland, NJ (Cumberland Co)", "feedback@usda.gov", "(856) 205-1225", "farmers.gov", "Apify", "MEDIUM", "USDA office. FSA loans, veteran programs.", ext_subj(), ext_body())

# TIER 3: LOW - Drone operators
add("Saturn Skylens LLC", "Drone service", "Moorestown, NJ (Burlington Co)", "business@saturnskylens.com", "(267) 209-6044", "saturnskylens.com", "Apify + scraped", "LOW", "Potential pilot hire.", pilot_subj(), pilot_body("Saturn Skylens"))
add("Adams Drone Service", "Drone service", "Burlington, NJ", "anthonygadams909@gmail.com", "(609) 556-9420", "adamsdroneservices.weebly.com", "Apify + scraped", "LOW", "Potential pilot hire.", pilot_subj(), pilot_body("Adams Drone Service"))
add("Precision Sky Drone Services", "Drone services", "Burlington County, NJ", "precisionskydroneservicesllc@gmail.com", "(609) 713-6316", "precisionskyrealestateanddroneservices.com", "Apify", "LOW", "Potential pilot hire.", pilot_subj(), pilot_body("Precision Sky"))
add("Blue Harbor Drones, LLC", "Drone service", "Cumberland County, NJ", "contact@blueharbordrones.com", "(856) 554-2319", "blueharbordrones.com", "Apify", "LOW", "Cumberland County operator.", pilot_subj(), pilot_body("Blue Harbor Drones"))
add("Drone Force", "Drone-based services", "Burlington County, NJ", "info@dronforcetech.com", "(856) 773-6723", "droneforcetech.com", "Apify", "LOW", "Potential hire.", pilot_subj(), pilot_body("Drone Force"))

# TIER 3: LOW - Farm supply
add("Fox Chase Feed & Supply", "Farm supply", "Southampton, NJ (Burlington Co)", "info@foxchasefeedandsupply.com", "", "foxchasefeedandsupply.com", "Apify", "LOW", "Farm supply Burlington Co.", supply_subj(), supply_body("Fox Chase Feed & Supply"))
add("Cherry Valley Tractor Sales", "Tractor dealer", "Marlton, NJ (Burlington Co)", "Keith@cherryvalleytractor.com", "(856) 983-0111", "cherryvalleytractor.com", "Apify", "LOW", "Knows every farmer.", supply_subj(), supply_body("Cherry Valley Tractor"))
add("Noble Turf", "Fertilizer supplier", "Mt Laurel, NJ (Burlington Co)", "noble@nobleturf.com", "(856) 273-5939", "nobleturf.com", "Apify", "LOW", "Fertilizer supplier.", supply_subj(), supply_body("Noble Turf"))
add("Stokes Seeds", "Seed supplier", "Vineland, NJ (Cumberland Co)", "custservice@stokeseeds.com", "(856) 692-6218", "stokeseeds.com", "Apify", "LOW", "Seed supplier.", supply_subj(), supply_body("Stokes Seeds"))

# TIER 3: LOW - Farm customers
add("J G Akerboom Nurseries Inc", "Agricultural service", "Cedarville, NJ (Cumberland Co)", "AKERBOOMSALES@HOTMAIL.COM", "(856) 447-3346", "akerboom.com", "Apify", "LOW", "Nursery/ag service.", farm_subj("Akerboom Nurseries"), farm_body("Akerboom Nurseries", "Cumberland County"))
add("Wilk Farm", "Farm", "Tabernacle, NJ (Burlington Co)", "mike@wilkfarm.com", "", "wilkfarm.com", "Apify", "LOW", "Farm in target area.", farm_subj("Wilk Farm"), farm_body("Wilk Farm", "Burlington County"))
add("Jersey Acres Farm", "Farm", "Medford, NJ (Burlington Co)", "erin@jerseyacresfarm.com", "(856) 236-3873", "jerseyacresfarm.com", "Apify", "LOW", "Farm in Burlington Co.", farm_subj("Jersey Acres Farm"), farm_body("Jersey Acres Farm", "Burlington County"))
add("Specca Farms", "Pick-your-own farm", "Bordentown, NJ (Burlington Co)", "speccafarms@gmail.com", "(609) 267-4445", "speccafarms.com", "Apify", "LOW", "Farm in Burlington Co.", farm_subj("Specca Farms"), farm_body("Specca Farms", "Burlington County"))
add("Honeycomb Homestead", "Agricultural service", "Medford Lakes, NJ (Burlington Co)", "savethebees@honeycombhomesteadnj.com", "(609) 234-9233", "honeycombhomesteadnj.com", "Apify", "LOW", "Ag service in Burlington Co.", farm_subj("Honeycomb Homestead"), farm_body("Honeycomb Homestead", "Burlington County"))
add("Whalen Farms", "Farm", "Shamong, NJ (Burlington Co)", "info@whalenfarms.com", "(609) 268-8047", "whalenfarms.com", "Apify", "LOW", "Established farm.", farm_subj("Whalen Farms"), farm_body("Whalen Farms", "Burlington County"))

# ── Build Excel ──
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Targets"

headers = list(targets[0].keys())
widths = [32, 28, 30, 30, 16, 38, 20, 10, 55, 45, 80]
for i, w in enumerate(widths):
    col_letter = chr(65 + i) if i < 26 else chr(64 + i // 26) + chr(65 + i % 26)
    ws.column_dimensions[col_letter].width = w

thin = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0")
)
hfill = PatternFill(start_color="102A43", end_color="102A43", fill_type="solid")
hfont = Font(bold=True, color="FFFFFF", size=10, name="Calibri")

for j, h in enumerate(headers):
    cell = ws.cell(row=1, column=j+1, value=h)
    cell.fill = hfill
    cell.font = hfont
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin

pcolors = {
    "HIGH": PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid"),
    "MEDIUM": PatternFill(start_color="FFF8F0", end_color="FFF8F0", fill_type="solid"),
    "LOW": PatternFill(start_color="F7F7F5", end_color="F7F7F5", fill_type="solid"),
}
pfont_map = {
    "HIGH": Font(bold=True, color="15803D", size=10, name="Calibri"),
    "MEDIUM": Font(bold=False, color="C45D1A", size=10, name="Calibri"),
    "LOW": Font(bold=False, color="7C7C72", size=10, name="Calibri"),
}

for i, t in enumerate(targets):
    pri = t["Priority"]
    for j, h in enumerate(headers):
        cell = ws.cell(row=i+2, column=j+1, value=t[h])
        cell.border = thin
        cell.font = Font(size=10, name="Calibri")
        cell.alignment = Alignment(wrap_text=(j >= 9), vertical="top")
        if j == 7:  # Priority
            cell.fill = pcolors.get(pri, pcolors["LOW"])
            cell.font = pfont_map.get(pri, pfont_map["LOW"])
            cell.alignment = Alignment(horizontal="center")

ws.auto_filter.ref = f"A1:K{len(targets)+1}"
ws.freeze_panes = "A2"

# Method sheet
ms = wb.create_sheet("Methodology")
ms.column_dimensions["A"].width = 80
notes = [
    "SOUTH JERSEY AG DRONE TARGETS - WITH EMAIL DRAFTS",
    "",
    f"Total targets: {len(targets)} (all with email + draft)",
    f"HIGH priority: {sum(1 for t in targets if t['Priority']=='HIGH')}",
    f"MEDIUM priority: {sum(1 for t in targets if t['Priority']=='MEDIUM')}",
    f"LOW priority: {sum(1 for t in targets if t['Priority']=='LOW')}",
    "",
    "Columns J & K contain pre-written email subject and body for each target.",
    "Each email is personalized by company name and tailored by outreach type:",
    "  - Acquisition: partnership/takeover language for drone competitors",
    "  - Competitor Intel: partnership framing for aerial applicators",
    "  - Pilot Recruitment: hiring pitch for drone operators",
    "  - Referral Partnership: mutual value prop for farm supply dealers",
    "  - Farm Customer: direct pitch for growers",
    "  - Extension/Government: introduction and demo offer",
    "",
    "All emails signed as Chris. Ready to copy-paste or load into CRM.",
]
for i, note in enumerate(notes):
    cell = ms.cell(row=i+1, column=1, value=note)
    if i == 0:
        cell.font = Font(bold=True, size=14, color="102A43", name="Calibri")
    else:
        cell.font = Font(size=10, name="Calibri")

wb.save("south_jersey_ag_drone_targets.xlsx")
print(f"Done: {len(targets)} targets with email drafts")
print(f"  HIGH: {sum(1 for t in targets if t['Priority']=='HIGH')}")
print(f"  MEDIUM: {sum(1 for t in targets if t['Priority']=='MEDIUM')}")
print(f"  LOW: {sum(1 for t in targets if t['Priority']=='LOW')}")
