import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

targets = []

def add(company, type_, location, email, phone, website, source, priority, notes):
    targets.append({
        "Company": company, "Type": type_, "Location": location,
        "Email": email, "Phone": phone, "Website": website,
        "Source": source, "Priority": priority, "Notes": notes,
    })

# TIER 1: Direct Competitors / Acquisition Targets
add("Influential Drones", "Ag drone spraying (Part 137)", "Marlton, NJ (Burlington Co)", "info@influentialdrones.com", "(856) 281-7545", "influentialdrones.com", "Apify + original", "HIGH", "Only confirmed NJ Part 137 drone sprayer. Top acquisition target.")
add("Alpha Drones USA", "Ag/commercial drone services", "Multi-state (NJ office)", "info@alphadronesusa.com", "1-800-270-4987", "alphadronesusa.com", "Perplexity + scraped", "HIGH", "Active NJ ag drone provider. Competitor or acquisition.")
add("Wings Aerial Applicators", "Traditional aerial applicator", "New Jersey", "wingsaerial@gmail.com", "(609) 760-5653", "wingsaerialapplicators.com", "Perplexity", "HIGH", "Incumbent manned aerial. 18+ years. Partnership or competitive intel.")
add("Downstown Aero", "Traditional aerial applicator", "Bridgeton, NJ (Cumberland Co)", "FireCats1@aol.com", "(856) 697-3300", "", "Perplexity", "HIGH", "Founded 1945. Fleet of 13 Ag-Cats. NJ prime aerial contractor since 1969.")
add("Burlington County Agricultural Center", "Farmers market / ag center", "Moorestown, NJ (Burlington Co)", "farmmarket@co.burlington.nj.us", "(856) 642-3850", "burlcoagcenter.com", "Apify", "HIGH", "County ag center. Perfect venue for demos and farmer networking.")
add("Rutgers Extension - Burlington County", "University extension", "Westampton, NJ", "bamka@njaes.rutgers.edu", "(609) 265-5050", "burlington.njaes.rutgers.edu", "Perplexity", "HIGH", "Top referral source. Gateway to Burlington Co blueberry farmers.")

# TIER 2: Competitors & Adjacent Services
add("AcuSpray", "Drone spraying technology", "National (NJ service)", "info@acuspray.com", "", "acuspray.com", "Perplexity + scraped", "MEDIUM", "Precision drone spraying company. Competitor or equipment partner.")
add("Delaware Valley Spray Services", "Lawn/ag spray service", "Hainesport, NJ (Burlington Co)", "info@delawarevalleysprayservice.com", "(609) 261-9400", "delawarevalleysprayservice.com", "Apify", "MEDIUM", "Spray services in Burlington Co. Potential competitor or partner.")
add("U.S.A.R. Drone Team (Agra Spray)", "Dealer / ag spray demos", "NJ/Northeast", "usardroneteam@gmail.com", "(732) 245-4234", "usardroneteam.org", "Original + scraped", "MEDIUM", "Regional ag spray operator/dealer network.")
add("NuWay Ag", "Dealer/pilot network", "New Jersey", "sales@nuwayag.com", "", "nuwayag.com", "Original + scraped", "MEDIUM", "NJ pilot map under construction. Key sourcing channel.")

# TIER 3: Local Drone Operators (potential hires/partners)
add("South Jersey Drone Services", "General drone services", "Sewell, NJ", "info@southjerseydroneservices.com", "(856) 807-1300", "southjerseydroneservices.com", "Original", "MEDIUM", "Local drone ops. Platform or talent source.")
add("National Drone Works", "Drone service", "Burlington County, NJ", "info@nationaldroneworks.com", "(856) 685-4188", "nationaldroneworks.com", "Apify", "MEDIUM", "Burlington County drone service. Potential pilot hire or partner.")
add("The Drone Life, LLC", "Aerial photography/drone", "Medford Lakes, NJ (Burlington Co)", "info@thedronelifenj.com", "(609) 500-5774", "thedronelifenj.com", "Apify", "MEDIUM", "Local drone operator. Potential pilot hire.")
add("Saturn Skylens LLC", "Drone service", "Moorestown, NJ (Burlington Co)", "business@saturnskylens.com", "(267) 209-6044", "saturnskylens.com", "Apify + scraped", "LOW", "Drone service in Burlington County. Potential pilot hire.")
add("Adams Drone Service", "Drone service", "Burlington, NJ", "anthonygadams909@gmail.com", "(609) 556-9420", "adamsdroneservices.weebly.com", "Apify + scraped", "LOW", "Burlington city drone service. Potential pilot hire.")
add("Precision Sky Drone Services", "Drone services", "Burlington County, NJ", "precisionskydroneservicesllc@gmail.com", "(609) 713-6316", "precisionskyrealestateanddroneservices.com", "Apify", "LOW", "Local drone operator. Potential pilot hire.")
add("Blue Harbor Drones, LLC", "Drone service", "Cumberland County, NJ", "contact@blueharbordrones.com", "(856) 554-2319", "blueharbordrones.com", "Apify", "LOW", "Cumberland County drone operator. Potential pilot hire.")
add("Drone Force", "Drone-based services", "Burlington County, NJ", "info@dronforcetech.com", "(856) 773-6723", "droneforcetech.com", "Apify", "LOW", "Drone services company. Potential hire.")

# TIER 4: Farm Supply Partners (referral sources)
add("Sussex Irrigation & Farm Supply", "Farm equipment supplier", "Bridgeton, NJ (Cumberland Co)", "info@sussexirrigation.com", "(856) 451-1368", "sussexirrigation.com", "Apify", "MEDIUM", "Farm supply dealer. Referral partner - knows every farmer.")
add("Burlington Agway", "Farm/feed supply", "Burlington, NJ", "burlingtonagway@live.com", "(609) 386-0500", "burlingtonagway.com", "Apify", "MEDIUM", "Feed/farm supply. Good referral channel to local growers.")
add("Ware's Farm Supply", "Farm supply", "Tabernacle, NJ (Burlington Co)", "waresfarmsupply@gmail.com", "(609) 268-1191", "waresfarmsupply.com", "Apify", "MEDIUM", "Farm supply in heart of blueberry country.")
add("Fox Chase Feed & Supply", "Farm supply", "Southampton, NJ (Burlington Co)", "info@foxchasefeedandsupply.com", "", "foxchasefeedandsupply.com", "Apify", "LOW", "Farm supply in Burlington Co.")
add("Cherry Valley Tractor Sales", "Tractor dealer", "Marlton, NJ (Burlington Co)", "Keith@cherryvalleytractor.com", "(856) 983-0111", "cherryvalleytractor.com", "Apify", "LOW", "Tractor dealer - knows every farmer in area.")
add("Noble Turf", "Fertilizer supplier", "Mt Laurel, NJ (Burlington Co)", "noble@nobleturf.com", "(856) 273-5939", "nobleturf.com", "Apify", "LOW", "Fertilizer supplier. Referral partner.")
add("Stokes Seeds", "Seed supplier", "Vineland, NJ (Cumberland Co)", "custservice@stokeseeds.com", "(856) 692-6218", "stokeseeds.com", "Apify", "LOW", "Seed supplier serving area growers.")

# TIER 5: Key Farms (potential customers)
add("Eastern Fresh Growers", "Vegetable wholesaler", "Cedarville, NJ (Cumberland Co)", "sales@easternfreshgrowers.com", "(856) 447-3563", "easternfreshgrowers.com", "Apify", "MEDIUM", "Major vegetable operation. Large potential customer.")
add("Oakland Farms Crop Services", "Farm / crop services", "Bridgeton, NJ (Cumberland Co)", "oaklandfarmsnj@gmail.com", "(856) 451-8224", "", "Apify", "MEDIUM", "Farm with crop services. Potential customer AND competitor.")
add("J G Akerboom Nurseries Inc", "Agricultural service", "Cedarville, NJ (Cumberland Co)", "AKERBOOMSALES@HOTMAIL.COM", "(856) 447-3346", "akerboom.com", "Apify", "LOW", "Nursery/ag service. Potential customer.")
add("Wilk Farm", "Farm", "Tabernacle, NJ (Burlington Co)", "mike@wilkfarm.com", "", "wilkfarm.com", "Apify", "LOW", "Farm in target area.")
add("Jersey Acres Farm", "Farm", "Medford, NJ (Burlington Co)", "erin@jerseyacresfarm.com", "(856) 236-3873", "jerseyacresfarm.com", "Apify", "LOW", "Farm in Burlington Co.")
add("Specca Farms", "Pick-your-own farm", "Bordentown, NJ (Burlington Co)", "speccafarms@gmail.com", "(609) 267-4445", "speccafarms.com", "Apify", "LOW", "Farm in Burlington Co.")
add("Honeycomb Homestead", "Agricultural service", "Medford Lakes, NJ (Burlington Co)", "savethebees@honeycombhomesteadnj.com", "(609) 234-9233", "honeycombhomesteadnj.com", "Apify", "LOW", "Ag service in Burlington Co.")
add("Whalen Farms", "Farm", "Shamong, NJ (Burlington Co)", "info@whalenfarms.com", "(609) 268-8047", "whalenfarms.com", "Apify", "LOW", "Established farm in target area.")

# TIER 6: Government
add("USDA Service Center - Vineland", "Federal ag office", "Vineland, NJ (Cumberland Co)", "feedback@usda.gov", "(856) 205-1225", "farmers.gov", "Apify", "MEDIUM", "USDA office. FSA loans, veteran programs, farmer directory.")

# ── Build Excel ──
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Targets"

headers = list(targets[0].keys())
widths = [35, 30, 32, 30, 18, 40, 22, 10, 60]
for i, w in enumerate(widths):
    ws.column_dimensions[chr(65+i)].width = w

thin = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0")
)
hfill = PatternFill(start_color="102A43", end_color="102A43", fill_type="solid")
hfont = Font(bold=True, color="FFFFFF", size=10, name="Calibri")

for j, h in enumerate(headers):
    cell = ws.cell(row=1, column=j+1, value=h)
    cell.fill = hfill
    cell.font = hfont
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin

pcolors = {
    "HIGH": PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid"),
    "MEDIUM": PatternFill(start_color="FFF8F0", end_color="FFF8F0", fill_type="solid"),
    "LOW": PatternFill(start_color="F7F7F5", end_color="F7F7F5", fill_type="solid"),
}
pfont_map = {
    "HIGH": Font(bold=True, color="15803D", size=10, name="Calibri"),
    "MEDIUM": Font(bold=False, color="C45D1A", size=10, name="Calibri"),
    "LOW": Font(bold=False, color="7C7C72", size=10, name="Calibri"),
}

for i, t in enumerate(targets):
    pri = t["Priority"]
    for j, h in enumerate(headers):
        cell = ws.cell(row=i+2, column=j+1, value=t[h])
        cell.border = thin
        cell.font = Font(size=10, name="Calibri")
        if j == 7:
            cell.fill = pcolors.get(pri, pcolors["LOW"])
            cell.font = pfont_map.get(pri, pfont_map["LOW"])
            cell.alignment = Alignment(horizontal="center")

ws.auto_filter.ref = f"A1:I{len(targets)+1}"
ws.freeze_panes = "A2"

# Method sheet
ms = wb.create_sheet("Methodology")
ms.column_dimensions["A"].width = 80
notes = [
    "SOUTH JERSEY AG DRONE TARGETS - CLEANED LIST",
    "",
    f"Total targets: {len(targets)} (all with email addresses)",
    f"HIGH priority: {sum(1 for t in targets if t['Priority']=='HIGH')}",
    f"MEDIUM priority: {sum(1 for t in targets if t['Priority']=='MEDIUM')}",
    f"LOW priority: {sum(1 for t in targets if t['Priority']=='LOW')}",
    "",
    "REMOVED: National brands (Tractor Supply, FairLifts, NAAA, Agri Spray directory, Peters Chemical)",
    "REMOVED: Entries with no findable email address",
    "ADDED: Scraped emails from company websites where missing",
    "",
    "SOURCES:",
    "  - Apify Google Maps scraping (Burlington + Cumberland counties)",
    "  - Perplexity AI deep research (3 queries)",
    "  - Website scraping for missing emails",
    "  - Original manual research targets",
]
for i, note in enumerate(notes):
    cell = ms.cell(row=i+1, column=1, value=note)
    if i == 0:
        cell.font = Font(bold=True, size=14, color="102A43", name="Calibri")
    elif note.startswith("Total") or note.startswith("HIGH") or note.startswith("MEDIUM") or note.startswith("LOW"):
        cell.font = Font(bold=True, size=10, name="Calibri")
    else:
        cell.font = Font(size=10, name="Calibri")

wb.save("south_jersey_ag_drone_targets.xlsx")
print(f"Done: {len(targets)} targets, all with emails")
print(f"  HIGH: {sum(1 for t in targets if t['Priority']=='HIGH')}")
print(f"  MEDIUM: {sum(1 for t in targets if t['Priority']=='MEDIUM')}")
print(f"  LOW: {sum(1 for t in targets if t['Priority']=='LOW')}")
