import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

targets = []

def add(company, type_, location, email, phone, website, source, priority, notes, biz_type):
    subject = f"Partnership interest in {company}"
    body = f"""Hi,

We're looking to partner with an owner of a {biz_type} in New Jersey and came across {company}.

Rather than a full sale upfront, we're interested in investing alongside an owner with a path to eventually taking over when they're ready to step back. The goal would be to support growth while keeping what's already working (your team, brand, and customer relationships).

I know this is out of the blue, but would you be open to a quick 10-15 minute intro call sometime?

If now's not the right time, I'd be happy to stay in touch down the road.

Best,
Chris"""
    targets.append({
        "Company": company, "Type": type_, "Location": location,
        "Email": email, "Phone": phone, "Website": website,
        "Source": source, "Priority": priority, "Notes": notes,
        "Email Subject": subject, "Email Body": body,
    })

# TIER 1: HIGH
add("Influential Drones", "Ag drone spraying (Part 137)", "Marlton, NJ (Burlington Co)", "info@influentialdrones.com", "(856) 281-7545", "influentialdrones.com", "Apify + original", "HIGH", "Only confirmed NJ Part 137 drone sprayer. Top acquisition target.", "drone spraying business")
add("Alpha Drones USA", "Ag/commercial drone services", "Multi-state (NJ office)", "info@alphadronesusa.com", "1-800-270-4987", "alphadronesusa.com", "Perplexity + scraped", "HIGH", "Active NJ ag drone provider. Competitor or acquisition.", "agricultural drone services business")
add("Wings Aerial Applicators", "Traditional aerial applicator", "New Jersey", "wingsaerial@gmail.com", "(609) 760-5653", "wingsaerialapplicators.com", "Perplexity", "HIGH", "Incumbent manned aerial. 18+ years.", "aerial application business")
add("Downstown Aero", "Traditional aerial applicator", "Bridgeton, NJ (Cumberland Co)", "FireCats1@aol.com", "(856) 697-3300", "", "Perplexity", "HIGH", "Founded 1945. Fleet of 13 Ag-Cats.", "agricultural aviation business")
add("Burlington County Agricultural Center", "Farmers market / ag center", "Moorestown, NJ (Burlington Co)", "farmmarket@co.burlington.nj.us", "(856) 642-3850", "burlcoagcenter.com", "Apify", "HIGH", "County ag center. Demo venue.", "agricultural center")
add("Rutgers Extension - Burlington County", "University extension", "Westampton, NJ", "bamka@njaes.rutgers.edu", "(609) 265-5050", "burlington.njaes.rutgers.edu", "Perplexity", "HIGH", "Gateway to Burlington Co blueberry farmers.", "agricultural extension program")

# TIER 2: MEDIUM - Competitors & Adjacent
add("AcuSpray", "Drone spraying technology", "National (NJ service)", "info@acuspray.com", "", "acuspray.com", "Perplexity + scraped", "MEDIUM", "Precision drone spraying. Competitor or partner.", "drone spraying business")
add("Delaware Valley Spray Services", "Lawn/ag spray service", "Hainesport, NJ (Burlington Co)", "info@delawarevalleysprayservice.com", "(609) 261-9400", "delawarevalleysprayservice.com", "Apify", "MEDIUM", "Spray services in Burlington Co.", "spray services business")
add("U.S.A.R. Drone Team (Agra Spray)", "Dealer / ag spray demos", "NJ/Northeast", "usardroneteam@gmail.com", "(732) 245-4234", "usardroneteam.org", "Original + scraped", "MEDIUM", "Regional ag spray operator/dealer.", "agricultural drone business")
add("NuWay Ag", "Dealer/pilot network", "New Jersey", "sales@nuwayag.com", "", "nuwayag.com", "Original + scraped", "MEDIUM", "NJ pilot map. Key sourcing channel.", "agricultural drone services business")

# TIER 2: MEDIUM - Drone operators
add("South Jersey Drone Services", "General drone services", "Sewell, NJ", "info@southjerseydroneservices.com", "(856) 807-1300", "southjerseydroneservices.com", "Original", "MEDIUM", "Local drone ops. Talent source.", "drone services business")
add("National Drone Works", "Drone service", "Burlington County, NJ", "info@nationaldroneworks.com", "(856) 685-4188", "nationaldroneworks.com", "Apify", "MEDIUM", "Burlington County drone service.", "drone services business")
add("The Drone Life, LLC", "Aerial photography/drone", "Medford Lakes, NJ (Burlington Co)", "info@thedronelifenj.com", "(609) 500-5774", "thedronelifenj.com", "Apify", "MEDIUM", "Local drone operator.", "drone services business")

# TIER 2: MEDIUM - Farm supply
add("Sussex Irrigation & Farm Supply", "Farm equipment supplier", "Bridgeton, NJ (Cumberland Co)", "info@sussexirrigation.com", "(856) 451-1368", "sussexirrigation.com", "Apify", "MEDIUM", "Farm supply. Referral partner.", "farm supply business")
add("Burlington Agway", "Farm/feed supply", "Burlington, NJ", "burlingtonagway@live.com", "(609) 386-0500", "burlingtonagway.com", "Apify", "MEDIUM", "Feed/farm supply. Referral channel.", "farm supply business")
add("Ware's Farm Supply", "Farm supply", "Tabernacle, NJ (Burlington Co)", "waresfarmsupply@gmail.com", "(609) 268-1191", "waresfarmsupply.com", "Apify", "MEDIUM", "Heart of blueberry country.", "farm supply business")

# TIER 2: MEDIUM - Farm customers
add("Eastern Fresh Growers", "Vegetable wholesaler", "Cedarville, NJ (Cumberland Co)", "sales@easternfreshgrowers.com", "(856) 447-3563", "easternfreshgrowers.com", "Apify", "MEDIUM", "Major vegetable operation.", "agricultural operation")
add("Oakland Farms Crop Services", "Farm / crop services", "Bridgeton, NJ (Cumberland Co)", "oaklandfarmsnj@gmail.com", "(856) 451-8224", "", "Apify", "MEDIUM", "Farm with crop services.", "crop services business")
add("USDA Service Center - Vineland", "Federal ag office", "Vineland, NJ (Cumberland Co)", "feedback@usda.gov", "(856) 205-1225", "farmers.gov", "Apify", "MEDIUM", "USDA office. FSA loans, veteran programs.", "agricultural services program")

# TIER 3: LOW - Drone operators
add("Saturn Skylens LLC", "Drone service", "Moorestown, NJ (Burlington Co)", "business@saturnskylens.com", "(267) 209-6044", "saturnskylens.com", "Apify + scraped", "LOW", "Potential pilot hire.", "drone services business")
add("Adams Drone Service", "Drone service", "Burlington, NJ", "anthonygadams909@gmail.com", "(609) 556-9420", "adamsdroneservices.weebly.com", "Apify + scraped", "LOW", "Potential pilot hire.", "drone services business")
add("Precision Sky Drone Services", "Drone services", "Burlington County, NJ", "precisionskydroneservicesllc@gmail.com", "(609) 713-6316", "precisionskyrealestateanddroneservices.com", "Apify", "LOW", "Potential pilot hire.", "drone services business")
add("Blue Harbor Drones, LLC", "Drone service", "Cumberland County, NJ", "contact@blueharbordrones.com", "(856) 554-2319", "blueharbordrones.com", "Apify", "LOW", "Cumberland County operator.", "drone services business")
add("Drone Force", "Drone-based services", "Burlington County, NJ", "info@dronforcetech.com", "(856) 773-6723", "droneforcetech.com", "Apify", "LOW", "Potential hire.", "drone services business")

# TIER 3: LOW - Farm supply
add("Fox Chase Feed & Supply", "Farm supply", "Southampton, NJ (Burlington Co)", "info@foxchasefeedandsupply.com", "", "foxchasefeedandsupply.com", "Apify", "LOW", "Farm supply Burlington Co.", "farm supply business")
add("Cherry Valley Tractor Sales", "Tractor dealer", "Marlton, NJ (Burlington Co)", "Keith@cherryvalleytractor.com", "(856) 983-0111", "cherryvalleytractor.com", "Apify", "LOW", "Knows every farmer.", "farm equipment business")
add("Noble Turf", "Fertilizer supplier", "Mt Laurel, NJ (Burlington Co)", "noble@nobleturf.com", "(856) 273-5939", "nobleturf.com", "Apify", "LOW", "Fertilizer supplier.", "turf and fertilizer business")
add("Stokes Seeds", "Seed supplier", "Vineland, NJ (Cumberland Co)", "custservice@stokeseeds.com", "(856) 692-6218", "stokeseeds.com", "Apify", "LOW", "Seed supplier.", "seed supply business")

# TIER 3: LOW - Farm customers
add("J G Akerboom Nurseries Inc", "Agricultural service", "Cedarville, NJ (Cumberland Co)", "AKERBOOMSALES@HOTMAIL.COM", "(856) 447-3346", "akerboom.com", "Apify", "LOW", "Nursery/ag service.", "nursery and agricultural business")
add("Wilk Farm", "Farm", "Tabernacle, NJ (Burlington Co)", "mike@wilkfarm.com", "", "wilkfarm.com", "Apify", "LOW", "Farm in target area.", "farming operation")
add("Jersey Acres Farm", "Farm", "Medford, NJ (Burlington Co)", "erin@jerseyacresfarm.com", "(856) 236-3873", "jerseyacresfarm.com", "Apify", "LOW", "Farm in Burlington Co.", "farming operation")
add("Specca Farms", "Pick-your-own farm", "Bordentown, NJ (Burlington Co)", "speccafarms@gmail.com", "(609) 267-4445", "speccafarms.com", "Apify", "LOW", "Farm in Burlington Co.", "farming operation")
add("Honeycomb Homestead", "Agricultural service", "Medford Lakes, NJ (Burlington Co)", "savethebees@honeycombhomesteadnj.com", "(609) 234-9233", "honeycombhomesteadnj.com", "Apify", "LOW", "Ag service in Burlington Co.", "agricultural business")
add("Whalen Farms", "Farm", "Shamong, NJ (Burlington Co)", "info@whalenfarms.com", "(609) 268-8047", "whalenfarms.com", "Apify", "LOW", "Established farm.", "farming operation")

# ── Build Excel ──
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Targets"

headers = list(targets[0].keys())
widths = [32, 28, 30, 30, 16, 38, 20, 10, 55, 40, 80]
for i, w in enumerate(widths):
    col_letter = chr(65 + i) if i < 26 else chr(64 + i // 26) + chr(65 + i % 26)
    ws.column_dimensions[col_letter].width = w

thin = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0")
)
hfill = PatternFill(start_color="102A43", end_color="102A43", fill_type="solid")
hfont = Font(bold=True, color="FFFFFF", size=10, name="Calibri")

for j, h in enumerate(headers):
    cell = ws.cell(row=1, column=j+1, value=h)
    cell.fill = hfill
    cell.font = hfont
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin

pcolors = {
    "HIGH": PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid"),
    "MEDIUM": PatternFill(start_color="FFF8F0", end_color="FFF8F0", fill_type="solid"),
    "LOW": PatternFill(start_color="F7F7F5", end_color="F7F7F5", fill_type="solid"),
}
pfont_map = {
    "HIGH": Font(bold=True, color="15803D", size=10, name="Calibri"),
    "MEDIUM": Font(bold=False, color="C45D1A", size=10, name="Calibri"),
    "LOW": Font(bold=False, color="7C7C72", size=10, name="Calibri"),
}

for i, t in enumerate(targets):
    pri = t["Priority"]
    for j, h in enumerate(headers):
        cell = ws.cell(row=i+2, column=j+1, value=t[h])
        cell.border = thin
        cell.font = Font(size=10, name="Calibri")
        cell.alignment = Alignment(wrap_text=(j >= 9), vertical="top")
        if j == 7:
            cell.fill = pcolors.get(pri, pcolors["LOW"])
            cell.font = pfont_map.get(pri, pfont_map["LOW"])
            cell.alignment = Alignment(horizontal="center")

ws.auto_filter.ref = f"A1:K{len(targets)+1}"
ws.freeze_panes = "A2"

wb.save("south_jersey_ag_drone_targets.xlsx")
print(f"Done: {len(targets)} targets, all with identical template emails")
print(f"  HIGH: {sum(1 for t in targets if t['Priority']=='HIGH')}")
print(f"  MEDIUM: {sum(1 for t in targets if t['Priority']=='MEDIUM')}")
print(f"  LOW: {sum(1 for t in targets if t['Priority']=='LOW')}")
