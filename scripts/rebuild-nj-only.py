import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

acq = []
refs = []
custs = []

def make_email(company, biz_type):
    subject = f"Partnership interest in {company}"
    body = f"""Hi,

We're looking to partner with an owner of a {biz_type} in New Jersey and came across {company}.

Rather than a full sale upfront, we're interested in investing alongside an owner with a path to eventually taking over when they're ready to step back. The goal would be to support growth while keeping what's already working (your team, brand, and customer relationships).

I know this is out of the blue, but would you be open to a quick 10-15 minute intro call sometime?

If now's not the right time, I'd be happy to stay in touch down the road.

Best,
Chris"""
    return subject, body

def add_acq(company, type_, location, email, phone, website, source, priority, notes, biz_type):
    subj, body = make_email(company, biz_type)
    acq.append({
        "Company": company, "Type": type_, "Location": location,
        "Email": email, "Phone": phone, "Website": website,
        "Source": source, "Priority": priority, "Notes": notes,
        "Email Subject": subj, "Email Body": body,
    })

def add_ref(company, type_, location, email, phone, website, source, priority, notes):
    refs.append({
        "Company": company, "Type": type_, "Location": location,
        "Email": email, "Phone": phone, "Website": website,
        "Source": source, "Priority": priority, "Notes": notes,
    })

def add_cust(company, type_, location, email, phone, website, source, priority, notes):
    custs.append({
        "Company": company, "Type": type_, "Location": location,
        "Email": email, "Phone": phone, "Website": website,
        "Source": source, "Priority": priority, "Notes": notes,
    })

# ═══════════════════════════════════════
# NJ ACQUISITION TARGETS ONLY
# ═══════════════════════════════════════

# HIGH - Ag drone / aerial spraying (direct competitors)
add_acq("Influential Drones", "Ag drone spraying (Part 137)", "Marlton, NJ (Burlington Co)", "info@influentialdrones.com", "(856) 281-7545", "influentialdrones.com", "Apify", "HIGH", "Only confirmed NJ Part 137 drone sprayer. #1 acquisition target.", "drone spraying business")
add_acq("Alpha Drones USA", "Ag/commercial drone services", "NJ (multi-state)", "info@alphadronesusa.com", "1-800-270-4987", "alphadronesusa.com", "Perplexity", "HIGH", "Active NJ ag drone provider. Major competitor.", "agricultural drone services business")
add_acq("Wings Aerial Applicators", "Traditional aerial applicator", "South NJ", "wingsaerial@gmail.com", "(609) 760-5653", "wingsaerialapplicators.com", "Perplexity", "HIGH", "Incumbent manned aerial. 18+ years. Family-owned.", "aerial application business")
add_acq("Downstown Aero", "Traditional aerial applicator", "Bridgeton, NJ (Cumberland Co)", "FireCats1@aol.com", "(856) 697-3300", "", "Perplexity", "HIGH", "Founded 1945. 13 Ag-Cats. NJ prime contractor since 1969.", "agricultural aviation business")
add_acq("Hawkeye Spray Drone Applications", "Ag drone spraying", "Annandale, NJ", "hawkeyespraydrone@gmail.com", "(908) 528-1993", "", "Perplexity", "HIGH", "NJ-based spray drone operator. Direct acquisition target.", "drone spraying business")
add_acq("Jersey Devil Dusters LLC", "Aerial applicator", "New Lisbon, NJ", "jerseydevildusters@gmail.com", "", "", "Perplexity", "HIGH", "NJ aerial applicator. Acquisition or partnership target.", "aerial application business")

# MEDIUM - NJ drone service companies (platform / talent)
add_acq("South Jersey Drone Services", "General drone services", "Sewell, NJ", "info@southjerseydroneservices.com", "(856) 807-1300", "southjerseydroneservices.com", "Original", "MEDIUM", "Local drone ops in South Jersey.", "drone services business")
add_acq("National Drone Works", "Drone service", "Burlington County, NJ", "info@nationaldroneworks.com", "(856) 685-4188", "nationaldroneworks.com", "Apify", "MEDIUM", "Burlington County drone service.", "drone services business")
add_acq("The Drone Life, LLC", "Drone service", "Medford Lakes, NJ (Burlington Co)", "info@thedronelifenj.com", "(609) 500-5774", "thedronelifenj.com", "Apify", "MEDIUM", "Local drone operator in target area.", "drone services business")
add_acq("U.S.A.R. Drone Team (Agra Spray)", "Ag drone dealer/operator", "NJ/Northeast", "usardroneteam@gmail.com", "(732) 245-4234", "usardroneteam.org", "Original", "MEDIUM", "Regional ag spray operator and DJI dealer.", "agricultural drone business")
add_acq("NuWay Ag", "Ag drone dealer/network", "New Jersey", "sales@nuwayag.com", "", "nuwayag.com", "Original", "MEDIUM", "NJ pilot network. Key operator channel.", "agricultural drone services business")
add_acq("Delaware Valley Spray Services", "Spray service", "Hainesport, NJ (Burlington Co)", "info@delawarevalleysprayservice.com", "(609) 261-9400", "delawarevalleysprayservice.com", "Apify", "MEDIUM", "Spray services in Burlington County.", "spray services business")
add_acq("AcuSpray", "Drone spraying technology", "NJ service area", "info@acuspray.com", "", "acuspray.com", "Perplexity", "MEDIUM", "Precision drone spraying company.", "drone spraying business")
add_acq("Volatus / Connexicore", "Commercial drone services", "NJ (nationwide)", "info@connexicore.com", "(855) 330-1780", "connexicore.com", "Perplexity", "MEDIUM", "FAA-certified drone services covering all NJ.", "drone services business")
add_acq("Osprey Perspectives", "Drone services", "NJ", "info@ospreyperspectives.com", "(732) 597-7461", "ospreyperspectives.com", "Perplexity", "MEDIUM", "FAA Part 107 with advanced waivers. NJ-based.", "drone services business")
add_acq("ABJ Drones", "Drone training/services", "Cranbury, NJ", "contact@abj.global", "(888) 225-1931", "abjdrones.com", "Apify", "MEDIUM", "Drone academy + services. Training pipeline.", "drone services and training business")
add_acq("Aerial View Advantage LLC", "Drone service", "Williamstown, NJ", "aerialviewadvantage@gmail.com", "(856) 371-4121", "aerialviewadvantage.com", "Apify", "MEDIUM", "South Jersey drone operator.", "drone services business")
add_acq("Enterprise Drone Solutions", "Commercial drone tech", "Warren County, NJ", "sales@enterprisedronesolutions.com", "", "enterprisedronesolutions.com", "BizBuySell", "MEDIUM", "Listed for sale on BizBuySell. Active acquisition opportunity.", "drone technology business")

# LOW - Smaller NJ drone operators
add_acq("Saturn Skylens LLC", "Drone service", "Moorestown, NJ (Burlington Co)", "business@saturnskylens.com", "(267) 209-6044", "saturnskylens.com", "Apify", "LOW", "Drone service in Burlington County.", "drone services business")
add_acq("Adams Drone Service", "Drone service", "Burlington, NJ", "anthonygadams909@gmail.com", "(609) 556-9420", "adamsdroneservices.weebly.com", "Apify", "LOW", "Burlington city drone service.", "drone services business")
add_acq("Precision Sky Drone Services", "Drone services", "Burlington County, NJ", "precisionskydroneservicesllc@gmail.com", "(609) 713-6316", "", "Apify", "LOW", "Local drone operator.", "drone services business")
add_acq("Blue Harbor Drones, LLC", "Drone service", "Cumberland County, NJ", "contact@blueharbordrones.com", "(856) 554-2319", "blueharbordrones.com", "Apify", "LOW", "Cumberland County drone operator.", "drone services business")
add_acq("Drone Force", "Drone service", "Burlington County, NJ", "info@dronforcetech.com", "(856) 773-6723", "droneforcetech.com", "Apify", "LOW", "Drone services company.", "drone services business")
add_acq("RedSky Drone", "Drone service", "Edison, NJ", "redskyvisuals@gmail.com", "", "redskydrone.com", "Apify", "LOW", "NJ drone photography/services.", "drone services business")
add_acq("The Fly Guys Photography", "Drone service", "Red Bank, NJ", "strbojc@gmail.com", "(732) 284-0879", "flyguysdroneservice.com", "Apify", "LOW", "NJ drone service.", "drone services business")
add_acq("NJ Drone Photo & Video", "Drone service", "Lyndhurst, NJ", "dave@nj-drone.com", "(862) 621-5305", "nj-drone.com", "Apify", "LOW", "NJ drone photography.", "drone services business")
add_acq("Skyview Media Co.", "Drone service", "North Brunswick, NJ", "info@skyviewmediaco.com", "(732) 538-7485", "skyviewmediaco.com", "Apify", "LOW", "NJ drone service.", "drone services business")
add_acq("North Jersey Drone Shots", "Drone service", "Bernardsville, NJ", "northjerseydroneshots@gmail.com", "(908) 507-5989", "northjerseydroneshots.com", "Apify", "LOW", "North Jersey drone operator.", "drone services business")
add_acq("Robinson Aerial Surveys", "Aerial services", "Hackettstown, NJ", "contactras@robinsonaerial.com", "(908) 813-3900", "robinsonaerial.com", "Apify", "LOW", "Established NJ aerial company.", "aerial services business")
add_acq("Aerial Rise", "Drone/aerial services", "Paterson, NJ", "info@aerialrise.com", "(973) 575-7480", "aerialrise.com", "Apify", "LOW", "NJ drone/aerial services.", "drone services business")
add_acq("Drone Services NJ/NY (uFly)", "Drone services", "North Bergen, NJ", "info@uflyuas.com", "", "uflyuas.com", "Apify", "LOW", "NJ/NY drone services.", "drone services business")

# ═══════════════════════════════════════
# REFERRAL PARTNERS (separate tab, no emails drafted)
# ═══════════════════════════════════════
add_ref("Rutgers Extension - Burlington County", "University extension", "Westampton, NJ", "bamka@njaes.rutgers.edu", "(609) 265-5050", "burlington.njaes.rutgers.edu", "Perplexity", "HIGH", "Gateway to Burlington Co blueberry farmers.")
add_ref("Burlington County Agricultural Center", "Ag center", "Moorestown, NJ", "farmmarket@co.burlington.nj.us", "(856) 642-3850", "burlcoagcenter.com", "Apify", "HIGH", "Perfect demo venue.")
add_ref("Sussex Irrigation & Farm Supply", "Farm supply", "Bridgeton, NJ", "info@sussexirrigation.com", "(856) 451-1368", "sussexirrigation.com", "Apify", "MEDIUM", "Farm supply. Referral partner.")
add_ref("Burlington Agway", "Farm supply", "Burlington, NJ", "burlingtonagway@live.com", "(609) 386-0500", "burlingtonagway.com", "Apify", "MEDIUM", "Feed/farm supply. Referral channel.")
add_ref("Ware's Farm Supply", "Farm supply", "Tabernacle, NJ", "waresfarmsupply@gmail.com", "(609) 268-1191", "waresfarmsupply.com", "Apify", "MEDIUM", "Heart of blueberry country.")
add_ref("USDA Service Center - Vineland", "Federal ag office", "Vineland, NJ", "feedback@usda.gov", "(856) 205-1225", "farmers.gov", "Apify", "MEDIUM", "FSA loans, veteran programs.")
add_ref("Fox Chase Feed & Supply", "Farm supply", "Southampton, NJ", "info@foxchasefeedandsupply.com", "", "foxchasefeedandsupply.com", "Apify", "LOW", "Farm supply Burlington Co.")
add_ref("Cherry Valley Tractor Sales", "Tractor dealer", "Marlton, NJ", "Keith@cherryvalleytractor.com", "(856) 983-0111", "cherryvalleytractor.com", "Apify", "LOW", "Knows every farmer.")
add_ref("Noble Turf", "Fertilizer supplier", "Mt Laurel, NJ", "noble@nobleturf.com", "(856) 273-5939", "nobleturf.com", "Apify", "LOW", "Fertilizer supplier.")
add_ref("Stokes Seeds", "Seed supplier", "Vineland, NJ", "custservice@stokeseeds.com", "(856) 692-6218", "stokeseeds.com", "Apify", "LOW", "Seed supplier.")

# ═══════════════════════════════════════
# POTENTIAL CUSTOMERS (separate tab)
# ═══════════════════════════════════════
add_cust("Eastern Fresh Growers", "Vegetable wholesaler", "Cedarville, NJ", "sales@easternfreshgrowers.com", "(856) 447-3563", "easternfreshgrowers.com", "Apify", "MEDIUM", "Major vegetable operation.")
add_cust("Oakland Farms Crop Services", "Crop services", "Bridgeton, NJ", "oaklandfarmsnj@gmail.com", "(856) 451-8224", "", "Apify", "MEDIUM", "Farm with crop services.")
add_cust("DiMeo Blueberry Farms", "Blueberry farm/nursery", "Hammonton, NJ", "dimeofarms@gmail.com", "(609) 561-5905", "dimeofarms.com", "Apify", "MEDIUM", "Major blueberry operation in Hammonton.")
add_cust("J G Akerboom Nurseries", "Nursery/ag", "Cedarville, NJ", "AKERBOOMSALES@HOTMAIL.COM", "(856) 447-3346", "akerboom.com", "Apify", "LOW", "Nursery/ag service.")
add_cust("Wilk Farm", "Farm", "Tabernacle, NJ", "mike@wilkfarm.com", "", "wilkfarm.com", "Apify", "LOW", "Farm in target area.")
add_cust("Jersey Acres Farm", "Farm", "Medford, NJ", "erin@jerseyacresfarm.com", "(856) 236-3873", "jerseyacresfarm.com", "Apify", "LOW", "Farm in Burlington Co.")
add_cust("Specca Farms", "Farm", "Bordentown, NJ", "speccafarms@gmail.com", "(609) 267-4445", "speccafarms.com", "Apify", "LOW", "Farm in Burlington Co.")
add_cust("Whalen Farms", "Farm", "Shamong, NJ", "info@whalenfarms.com", "(609) 268-8047", "whalenfarms.com", "Apify", "LOW", "Established farm.")
add_cust("Ward's Farm NJ", "Farm", "Mannington, NJ", "WardsFarmNJ@gmail.com", "(856) 279-0404", "wardsfarmnj.com", "Apify", "LOW", "Salem County farm.")
add_cust("Ferrara's Orchards", "Farm/garden center", "Mullica Hill, NJ", "ferrarasorchards@gmail.com", "(856) 478-0208", "ferrarasorchards.com", "Apify", "LOW", "Gloucester County orchard.")

# ═══════════════════════════════════════
# BUILD EXCEL
# ═══════════════════════════════════════
wb = openpyxl.Workbook()

thin = Border(
    left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0")
)
hfill = PatternFill(start_color="102A43", end_color="102A43", fill_type="solid")
hfont = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
pcolors = {
    "HIGH": PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid"),
    "MEDIUM": PatternFill(start_color="FFF8F0", end_color="FFF8F0", fill_type="solid"),
    "LOW": PatternFill(start_color="F7F7F5", end_color="F7F7F5", fill_type="solid"),
}
pfont_map = {
    "HIGH": Font(bold=True, color="15803D", size=10, name="Calibri"),
    "MEDIUM": Font(bold=False, color="C45D1A", size=10, name="Calibri"),
    "LOW": Font(bold=False, color="7C7C72", size=10, name="Calibri"),
}

def build_sheet(wb, title, data, tab_color, widths_list, is_first=False):
    ws = wb.active if is_first else wb.create_sheet(title)
    ws.title = title
    ws.sheet_properties.tabColor = tab_color
    headers = list(data[0].keys())
    for i, w in enumerate(widths_list):
        col = chr(65+i) if i < 26 else chr(64+i//26) + chr(65+i%26)
        ws.column_dimensions[col].width = w
    for j, h in enumerate(headers):
        cell = ws.cell(row=1, column=j+1, value=h)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center", wrap_text=True); cell.border = thin
    for i, t in enumerate(data):
        pri = t.get("Priority", "LOW")
        for j, h in enumerate(headers):
            cell = ws.cell(row=i+2, column=j+1, value=t[h])
            cell.border = thin; cell.font = Font(size=10, name="Calibri")
            cell.alignment = Alignment(wrap_text=(j >= 9), vertical="top")
            if h == "Priority":
                cell.fill = pcolors.get(pri, pcolors["LOW"])
                cell.font = pfont_map.get(pri, pfont_map["LOW"])
                cell.alignment = Alignment(horizontal="center")
    ws.auto_filter.ref = f"A1:{chr(64+len(headers))}{len(data)+1}"
    ws.freeze_panes = "A2"

build_sheet(wb, "Acquisition Targets", acq, "16A34A",
    [32, 28, 26, 30, 16, 35, 18, 10, 55, 40, 80], is_first=True)
build_sheet(wb, "Referral Partners", refs, "E8792A",
    [32, 22, 24, 30, 16, 35, 18, 10, 55])
build_sheet(wb, "Potential Customers", custs, "486581",
    [32, 22, 24, 30, 16, 35, 18, 10, 55])

wb.save("south_jersey_ag_drone_targets.xlsx")
print(f"Acquisition Targets: {len(acq)} (all NJ, all with email drafts)")
print(f"  HIGH: {sum(1 for t in acq if t['Priority']=='HIGH')}")
print(f"  MEDIUM: {sum(1 for t in acq if t['Priority']=='MEDIUM')}")
print(f"  LOW: {sum(1 for t in acq if t['Priority']=='LOW')}")
print(f"Referral Partners: {len(refs)}")
print(f"Potential Customers: {len(custs)}")
print(f"TOTAL: {len(acq) + len(refs) + len(custs)}")
